import pandas as pd
import sqlite3
import os
from openpyxl import load_workbook

# --- CONFIGURAÇÃO ---
DB_PATH = 'vendas_animoshop.db'
EXCEL_PATH = 'Balanço Out-25.xlsx'
ANO = 2025
ANO = 2025
MES = 'Outubro' # Database uses month names

# Mapeamento Linha Excel -> Nome Marketplace no Banco
# Apenas os que temos no banco
MAPA_LINHAS = {
    19: 'Amazon',
    24: 'MadeiraMadeira',
    25: 'Magalu',
    26: 'Mercado Livre',
    27: 'Olist',
    28: 'Shopee'
}

def get_db_totals():
    """Busca totais de Comissões e Frete por Marketplace no Banco."""
    conn = sqlite3.connect(DB_PATH)
    
    # Query unificando todas as tabelas (Limpas? ou Consolidado Geral?)
    # Se existir 'animoshop_consolidado_geral', melhor usar ela.
    # Mas o nome certo costuma ser 'shopee_consolidado', 'amazon_consolidado', etc.
    # Vamos iterar pelas tabelas conhecidas.
    
    marketplaces_db = [
        'amazon_consolidado', 'shopee_consolidado', 'mercado_livre_consolidado',
        'magalu_consolidado', 'madeira_consolidado', 'olist_consolidado'
    ]
    
    resultados = {} # {'Amazon': {'comissao': 100, 'frete': 50}, ...}

    for table in marketplaces_db:
        try:
            # Check if table exists
            check = pd.read_sql_query(f"SELECT name FROM sqlite_master WHERE type='table' AND name='{table}'", conn)
            if check.empty:
                continue

            query = f"""
                SELECT 
                    SUM(comissoes) as comissao, 
                    SUM(frete) as frete 
                FROM {table}
                WHERE ano = {ANO} AND mes = '{MES}'
            """
            df = pd.read_sql_query(query, conn)
            
            # Debug se vazio
            if df['comissao'].isnull().all():
                 # Check available months
                 months_df = pd.read_sql_query(f"SELECT DISTINCT mes, ano FROM {table}", conn)
                 print(f"⚠ No data for {MES}/{ANO} in {table}. Available: {months_df.values.tolist()}")
            
            # Identify DB marketplace name from table name
            mp_name = table.replace('_consolidado', '').replace('_', ' ').title()
            # Ajustes manuais de nome para bater com MAPA_LINHAS
            if 'Madeira' in mp_name: mp_name = 'MadeiraMadeira'
            if 'Mercado Livre' in mp_name: mp_name = 'Mercado Livre'
            
            comissao = df['comissao'].fillna(0).values[0]
            frete = df['frete'].fillna(0).values[0]
            
            # Ensure positive values (sometimes expenses are negative in DB)
            resultados[mp_name] = {
                'comissao': abs(comissao),
                'frete': abs(frete)
            }
            print(f"Db Data for {mp_name}: Comissões={abs(comissao):.2f}, Frete={abs(frete):.2f}")

        except Exception as e:
            print(f"Error querying {table}: {e}")
            
    conn.close()
    return resultados

def update_excel(dados_db):
    """Atualiza a planilha Excel."""
    try:
        wb = load_workbook(EXCEL_PATH)
        ws = wb.active # Assume primeira aba

        print("\n--- Atualizando Excel ---")
        for row, mp_name in MAPA_LINHAS.items():
            if mp_name in dados_db:
                comissao = dados_db[mp_name]['comissao']
                frete = dados_db[mp_name]['frete']
                
                # Col X (24) = Comissões
                ws.cell(row=row, column=24, value=comissao)
                
                # Col Y (25) = Frete
                ws.cell(row=row, column=25, value=frete)
                
                print(f"Updated Row {row} ({mp_name}): Com={comissao:.2f}, Frete={frete:.2f}")
            else:
                print(f"Skipping Row {row} ({mp_name}) - No data in DB for Oct/25")

        wb.save(EXCEL_PATH)
        print(f"\n✅ Arquivo salvo com sucesso: {EXCEL_PATH}")

    except Exception as e:
        print(f"Erro ao atualizar Excel: {e}")

if __name__ == "__main__":
    print(f"Buscando dados para {MES}/{ANO}...")
    dados = get_db_totals()
    update_excel(dados)
