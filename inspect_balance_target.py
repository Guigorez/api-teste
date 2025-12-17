import pandas as pd
from openpyxl import load_workbook

file_path = 'Balanço Out-25.xlsx'

try:
    # Use openpyxl to inspect exact cells including headers if any nearby
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active # Assuming first sheet
    
    print(f"--- Content of W17:Y30 in {ws.title} ---")
    
    # Header row might be above 17? Let's check 16-17 to be sure
    # Rows are 1-based in openpyxl
    for i, row in enumerate(ws.iter_rows(min_row=16, max_row=30, min_col=23, max_col=25)): # W=23, Y=25
        row_num = 16 + i
        vals = [c.value for c in row]
        print(f"Row {row_num}: {vals}")

except Exception as e:
    print(f"Error: {e}")
